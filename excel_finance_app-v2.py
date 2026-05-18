"""
Finance Template Generator v2
Generates premium, sellable Excel spreadsheet templates.
Run: python excel_finance_app-v2.py
Requires: pip install openpyxl rich questionary
"""

import os, sys, datetime, pathlib, random
import questionary
from rich.console import Console
from rich.panel import Panel
from rich.columns import Columns
from rich.text import Text
from rich import box

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

console = Console()

# ─── Vibes ────────────────────────────────────────────────────────────────────

VIBES = {
    "Neutral Luxe": {
        "desc": "Warm beige, taupe & terracotta — timeless, broad appeal",
        "font": "Book Antiqua",
        "bg":         "F7F1E8",
        "header_bg":  "C9A97A",
        "accent":     "8B6347",
        "accent2":    "D4B896",
        "text":       "2C1810",
        "text_light": "FFFFFF",
        "subtext":    "8B7355",
        "alt_row":    "FBF7F1",
        "border":     "D4C4B0",
        "kpi_bg":     "EDE0CE",
        "stripe":     "C9A97A",
        "chart_style": 8,
    },
    "Dark & Minimal": {
        "desc": "Charcoal, slate & off-white — professional, sharp",
        "font": "Calibri Light",
        "bg":         "1A1F26",
        "header_bg":  "252C35",
        "accent":     "4FC3F7",
        "accent2":    "64FFDA",
        "text":       "E8EAED",
        "text_light": "E8EAED",
        "subtext":    "9AA5B4",
        "alt_row":    "1E252D",
        "border":     "3D4654",
        "kpi_bg":     "2A323D",
        "stripe":     "4FC3F7",
        "chart_style": 23,
    },
    "Soft Feminine": {
        "desc": "Blush, sage & lavender — popular with lifestyle audiences",
        "font": "Century Gothic",
        "bg":         "FFF5F7",
        "header_bg":  "F4ACB7",
        "accent":     "9D6B7A",
        "accent2":    "B8D8BA",
        "text":       "4A3540",
        "text_light": "FFFFFF",
        "subtext":    "9D8189",
        "alt_row":    "FFF0F3",
        "border":     "F4C0CC",
        "kpi_bg":     "FFE4E9",
        "stripe":     "F4ACB7",
        "chart_style": 12,
    },
    "Bold & Modern": {
        "desc": "Deep navy, clean white & gold — business owners, investors",
        "font": "Calibri",
        "bg":         "FAFBFC",
        "header_bg":  "1B263B",
        "accent":     "C9A84C",
        "accent2":    "415A77",
        "text":       "1B263B",
        "text_light": "FFFFFF",
        "subtext":    "64748B",
        "alt_row":    "F0F4F8",
        "border":     "CBD5E1",
        "kpi_bg":     "E8F0FE",
        "stripe":     "C9A84C",
        "chart_style": 2,
    },
}

# ─── Currencies ───────────────────────────────────────────────────────────────

CURRENCIES = {
    "GBP  £   British Pound":    ("£",  "#,##0.00"),
    "EUR  €   Euro":             ("€",  "#,##0.00"),
    "USD  $   US Dollar":        ("$",  "#,##0.00"),
    "HUF  Ft  Hungarian Forint": ("Ft", "#,##0"),
    "JPY  ¥   Japanese Yen":     ("¥",  "#,##0"),
    "CAD  C$  Canadian Dollar":  ("C$", "#,##0.00"),
    "AUD  A$  Australian Dollar":("A$", "#,##0.00"),
    "CHF      Swiss Franc":      ("CHF","#,##0.00"),
    "PLN  zł  Polish Zloty":     ("zł", "#,##0.00"),
    "CZK  Kč  Czech Koruna":     ("Kč", "#,##0"),
}

TEMPLATE_TYPES = {
    "Personal Budget Tracker":
        "Monthly income & expenses, spending breakdown, savings rate",
    "Freelancer / Self-Employed":
        "Income log, expense tracker, tax estimate, P&L summary",
    "Savings Goals Planner":
        "Multiple goals, progress tracking, contribution log",
}

OUTPUT_DIR = pathlib.Path(__file__).parent / "output"
LOG_FILE   = pathlib.Path(__file__).parent / "answers_log.txt"

# ─── Style helpers ────────────────────────────────────────────────────────────

def F(hex6): return PatternFill("solid", fgColor=hex6)
def Fn(hex6, bold=False, size=11, italic=False, name="Calibri"):
    return Font(color=hex6, bold=bold, size=size, italic=italic, name=name)
def Al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def Br(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def BrNone():
    n = Side(style=None)
    return Border(left=n, right=n, top=n, bottom=n)

def cell(ws, row, col, value=None, fill=None, font=None, align=None, border=None, fmt=None, height=None):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    if fill   is not None: c.fill   = fill
    if font   is not None: c.font   = font
    if align  is not None: c.alignment = align
    if border is not None: c.border = border
    if fmt    is not None: c.number_format = fmt
    if height is not None: ws.row_dimensions[row].height = height
    return c

def merge(ws, r1, c1, r2, c2, value=None, fill=None, font=None, align=None, border=None, height=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return cell(ws, r1, c1, value, fill, font, align, border, height=height)

def row_fill(ws, row, col_start, col_end, fill, height=None, border=None):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = fill
        if border: ws.cell(row=row, column=c).border = border
    if height: ws.row_dimensions[row].height = height

def table_header(ws, row, headers, vibe, font_name, col_start=1, height=22):
    v = vibe
    for i, h in enumerate(headers, col_start):
        cell(ws, row, i, h,
             fill=F(v["header_bg"]),
             font=Fn(v["text_light"], bold=True, size=9, name=font_name),
             align=Al("center", "center"),
             border=Br(v["border"]))
    ws.row_dimensions[row].height = height

def table_row(ws, row, values, vibe, font_name, col_start=1, alt=False, height=18, fmts=None):
    v = vibe
    bg = v["alt_row"] if alt else v["bg"]
    for i, val in enumerate(values, col_start):
        fmt = fmts[i - col_start] if fmts else None
        cell(ws, row, i, val,
             fill=F(bg),
             font=Fn(v["text"], size=9, name=font_name),
             align=Al("center", "center"),
             border=Br(v["border"]),
             fmt=fmt)
    ws.row_dimensions[row].height = height

def kpi_box(ws, r1, c1, r2, c2, label, formula, vibe, font_name, num_fmt="#,##0.00"):
    v = vibe
    merge(ws, r1, c1, r1, c2, label,
          fill=F(v["header_bg"]),
          font=Fn(v["text_light"], bold=True, size=8, name=font_name),
          align=Al("center", "bottom"))
    merge(ws, r1+1, c1, r2, c2, formula,
          fill=F(v["kpi_bg"]),
          font=Fn(v["accent"], bold=True, size=18, name=font_name),
          align=Al("center", "center"))
    ws.cell(row=r1+1, column=c1).number_format = num_fmt

def section_bar(ws, row, col_start, col_end, label, vibe, font_name, height=18):
    v = vibe
    merge(ws, row, col_start, row, col_end, label,
          fill=F(v["accent"]),
          font=Fn(v["text_light"], bold=True, size=9, name=font_name),
          align=Al("left", "center"),
          height=height)
    # left padding via indent
    ws.cell(row=row, column=col_start).alignment = Alignment(
        horizontal="left", vertical="center", indent=1)

def col_widths(ws, widths, start=1):
    for i, w in enumerate(widths, start):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─── Sample data ──────────────────────────────────────────────────────────────

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]
MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun",
                "Jul","Aug","Sep","Oct","Nov","Dec"]

BUDGET_SAMPLE = [
    # income, housing, food, transport, utilities, entertainment, personal, health, other
    (2800, 900, 295, 92, 78, 135, 88, 45, 62),
    (2800, 900, 312, 88, 82, 110, 95, 0,  55),
    (3100, 900, 280, 95, 75, 150, 72, 60, 48),
    (2800, 900, 305, 91, 79, 125, 85, 0,  70),
    (2800, 900, 318, 86, 77, 160, 98, 30, 52),
    (3200, 900, 290, 98, 80, 145, 80, 0,  60),
    (2800, 900, 330, 102, 76, 200, 110, 45, 75),
    (2800, 900, 325, 94, 75, 185, 92, 0,  58),
    (2800, 900, 298, 89, 81, 140, 78, 40, 65),
    (3000, 900, 285, 88, 83, 120, 88, 0,  50),
    (2800, 900, 310, 95, 79, 130, 95, 55, 48),
    (3400, 900, 380, 110, 82, 280, 150, 0, 120),
]

FREELANCER_INCOME_SAMPLE = [
    ("05 Jan 2026","Bright Creative Ltd","Brand Identity Package",1800,"Paid"),
    ("12 Jan 2026","The Studio Co","Social Media Content Pack",650,"Paid"),
    ("20 Jan 2026","Oakwood Agency","Website Copy",950,"Paid"),
    ("03 Feb 2026","Bright Creative Ltd","Ongoing Retainer",1200,"Paid"),
    ("18 Feb 2026","Nova Digital","Logo & Brand Guide",1400,"Paid"),
    ("25 Feb 2026","Freelance Client","Photography Edit",400,"Paid"),
    ("04 Mar 2026","The Studio Co","Quarterly Campaign",2200,"Paid"),
    ("15 Mar 2026","Oakwood Agency","Consultation",350,"Paid"),
    ("22 Mar 2026","New Client A","Landing Page Design",1100,"Paid"),
    ("01 Apr 2026","Bright Creative Ltd","Ongoing Retainer",1200,"Paid"),
    ("10 Apr 2026","Nova Digital","Social Ads Pack",780,"Paid"),
    ("28 Apr 2026","Freelance Client","Video Script",500,"Unpaid"),
    ("05 May 2026","New Client B","Brand Refresh",1650,"Unpaid"),
    ("14 May 2026","The Studio Co","Email Campaign",600,"Unpaid"),
    ("","","","",""),
    ("","","","",""),
    ("","","","",""),
    ("","","","",""),
    ("","","","",""),
    ("","","","",""),
]

FREELANCER_EXPENSE_SAMPLE = [
    ("04 Jan 2026","Software","Adobe Creative Cloud",54.99),
    ("04 Jan 2026","Software","Notion Pro",8.00),
    ("10 Jan 2026","Equipment","External Hard Drive",79.00),
    ("04 Feb 2026","Software","Adobe Creative Cloud",54.99),
    ("04 Feb 2026","Software","Notion Pro",8.00),
    ("20 Feb 2026","Marketing","Instagram Ads",50.00),
    ("04 Mar 2026","Software","Adobe Creative Cloud",54.99),
    ("04 Mar 2026","Software","Notion Pro",8.00),
    ("15 Mar 2026","Travel","Client Meeting Travel",35.00),
    ("04 Apr 2026","Software","Adobe Creative Cloud",54.99),
    ("04 Apr 2026","Software","Notion Pro",8.00),
    ("18 Apr 2026","Professional Services","Accountant",120.00),
    ("04 May 2026","Software","Adobe Creative Cloud",54.99),
    ("04 May 2026","Software","Notion Pro",8.00),
    ("","","",None),
    ("","","",None),
    ("","","",None),
    ("","","",None),
    ("","","",None),
    ("","","",None),
]

GOALS_SAMPLE = [
    ("Emergency Fund",    10000, 4200, 500),
    ("Holiday — 2026",    3500,  1750, 350),
    ("New Laptop",        1800,  900,  150),
    ("House Deposit",     25000, 8500, 800),
    ("Car Fund",          12000, 2400, 400),
]

# ─── Dashboard builders ───────────────────────────────────────────────────────

def _dash_header(ws, vibe, font_name, title, subtitle, ncols=14):
    v = vibe
    # Top stripe
    row_fill(ws, 1, 1, ncols, F(v["stripe"]), height=6)
    # Logo area + Title
    ws.row_dimensions[2].height = 52
    # Logo placeholder (cols 1-3)
    merge(ws, 2, 1, 3, 3, "[ YOUR LOGO ]",
          fill=F(v["kpi_bg"]),
          font=Fn(v["subtext"], italic=True, size=8, name=font_name),
          align=Al("center", "center"))
    # Main title (cols 4-14)
    merge(ws, 2, 4, 2, ncols, title,
          fill=F(v["bg"]),
          font=Fn(v["accent"], bold=True, size=22, name=font_name),
          align=Al("left", "bottom"))
    ws.cell(row=2, column=4).alignment = Alignment(
        horizontal="left", vertical="bottom", indent=1)
    # Subtitle row
    ws.row_dimensions[3].height = 20
    merge(ws, 3, 4, 3, ncols, subtitle,
          fill=F(v["bg"]),
          font=Fn(v["subtext"], italic=True, size=9, name=font_name),
          align=Al("left", "top"))
    ws.cell(row=3, column=4).alignment = Alignment(
        horizontal="left", vertical="top", indent=1)
    # Bottom stripe
    row_fill(ws, 4, 1, ncols, F(v["accent"]), height=3)


def _currency_settings(ws, vibe, font_name, row, sym_default, ncols=14):
    v = vibe
    row_fill(ws, row, 1, ncols, F(v["bg"]), height=8)
    row_fill(ws, row+1, 1, ncols, F(v["bg"]), height=22)

    cell(ws, row+1, 1, "Currency Symbol:",
         fill=F(v["bg"]),
         font=Fn(v["subtext"], bold=True, size=9, name=font_name),
         align=Al("right", "center"))
    ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=3)
    c = ws.cell(row=row+1, column=2)
    c.value = sym_default
    c.fill  = F(v["kpi_bg"])
    c.font  = Fn(v["accent"], bold=True, size=11, name=font_name)
    c.alignment = Al("center", "center")
    c.border = Br(v["accent"])

    # Currency dropdown
    cur_symbols = "£,€,$,Ft,¥,C$,A$,CHF,zł,Kč"
    dv = DataValidation(type="list", formula1=f'"{cur_symbols}"', showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row+1, column=2))

    # Label explaining use
    merge(ws, row+1, 4, row+1, ncols,
          "← Change this cell to update currency labels throughout the dashboard.",
          fill=F(v["bg"]),
          font=Fn(v["subtext"], italic=True, size=8, name=font_name),
          align=Al("left", "center"))

    # Named range for currency symbol so dashboard formulas can reference it
    # We'll reference Dashboard!$B$<row+1> in formulas directly
    return row + 1  # returns the row of the currency cell


# ─── Personal Budget Template ─────────────────────────────────────────────────

def build_budget_template(vibe_name, sym, num_fmt, year=None):
    v    = VIBES[vibe_name]
    fn   = v["font"]
    ncols = 14
    if year is None: year = datetime.date.today().year
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet: Dashboard ──────────────────────────────────────────────────────
    dash = wb.create_sheet("Dashboard")
    dash.sheet_view.showGridLines = False
    dash.sheet_view.zoomScale = 90

    _dash_header(dash, v, fn,
                 f"PERSONAL BUDGET TRACKER  {year}",
                 "Your complete financial overview at a glance",
                 ncols)

    cur_row_num = _currency_settings(dash, v, fn, 5, sym, ncols)
    # cur_row_num = 6
    CUR_CELL = f"Dashboard!$B${cur_row_num}"

    # Spacer
    row_fill(dash, 7, 1, ncols, F(v["bg"]), height=10)

    # KPI section bar
    section_bar(dash, 8, 1, ncols, "  ANNUAL SUMMARY", v, fn, height=18)

    # KPI boxes: row 9 (label) rows 9-11 (value)
    dash.row_dimensions[9].height = 18
    dash.row_dimensions[10].height = 36
    dash.row_dimensions[11].height = 8

    budget_sheet = "Budget"
    kpi_data = [
        ("TOTAL INCOME",    f"=IFERROR(SUM('{budget_sheet}'!C7:C18),0)", 1, 3),
        ("TOTAL EXPENSES",  f"=IFERROR(SUM('{budget_sheet}'!L7:L18),0)", 4, 6),
        ("NET SAVINGS",     f"=IFERROR(SUM('{budget_sheet}'!M7:M18),0)", 7, 9),
        ("AVG SAVINGS RATE",f"=IFERROR(AVERAGE('{budget_sheet}'!N7:N18),0)", 10, 13),
    ]
    for label, formula, cs, ce in kpi_data:
        fmt = "0.0%" if "RATE" in label else num_fmt
        kpi_box(dash, 9, cs, 11, ce, label, formula, v, fn, num_fmt=fmt)
    # fill col 14 beside last KPI
    for r in range(9, 12):
        dash.cell(row=r, column=14).fill = F(v["bg"])

    # Spacer
    row_fill(dash, 12, 1, ncols, F(v["bg"]), height=10)

    # Charts section bar
    section_bar(dash, 13, 1, ncols, "  CHARTS & INSIGHTS", v, fn)

    # Reserve rows 14-38 for charts (openpyxl places them as floating objects)
    for r in range(14, 42):
        row_fill(dash, r, 1, ncols, F(v["bg"]), height=15)

    # Chart: monthly spending categories bar chart (data from Budget sheet)
    # We'll reference Budget rows 4-15, cols: A (month), C-K (income, cats), L (total exp)
    budget_ws_placeholder = None  # we build it next, chart uses sheet name reference

    # Monthly trend bar chart
    chart_bar = BarChart()
    chart_bar.type = "col"
    chart_bar.grouping = "clustered"
    chart_bar.title = "Monthly Income vs Expenses"
    chart_bar.style = v["chart_style"]
    chart_bar.y_axis.title = "Amount"
    chart_bar.x_axis.title = "Month"
    chart_bar.width = 22; chart_bar.height = 14
    chart_bar.legend.position = "b"

    # Donut chart
    chart_donut = DoughnutChart()
    chart_donut.title = "Spending Breakdown"
    chart_donut.style = v["chart_style"]
    chart_donut.width = 14; chart_donut.height = 14
    chart_donut.holeSize = 55
    chart_donut.legend.position = "r"

    # Spacer before monthly table
    row_fill(dash, 42, 1, ncols, F(v["bg"]), height=10)

    # Monthly summary table
    section_bar(dash, 43, 1, ncols, "  MONTHLY SUMMARY", v, fn)
    tbl_headers = ["Month", "Income", "Total Expenses", "Net Savings", "Savings Rate"]
    table_header(dash, 44, tbl_headers, v, fn, col_start=1)
    # col widths for these 5 cols: 12, 14, 16, 14, 14
    for mi, mname in enumerate(MONTHS_SHORT):
        r = 45 + mi
        row_data = [
            mname,
            f"=IFERROR('{budget_sheet}'!C{7+mi},0)",
            f"=IFERROR('{budget_sheet}'!L{7+mi},0)",
            f"=IFERROR('{budget_sheet}'!M{7+mi},0)",
            f"=IFERROR('{budget_sheet}'!N{7+mi},0)",
        ]
        fmts = [None, num_fmt, num_fmt, num_fmt, "0.0%"]
        table_row(dash, r, row_data, v, fn, col_start=1, alt=(mi%2==0), fmts=fmts)

    # Totals row
    table_row(dash, 57, [
        "TOTAL",
        f"=SUM(B45:B56)", f"=SUM(C45:C56)", f"=SUM(D45:D56)", f"=AVERAGE(E45:E56)"
    ], v, fn, col_start=1, fmts=[None, num_fmt, num_fmt, num_fmt, "0.0%"])
    for c in range(1, 6):
        dash.cell(row=57, column=c).fill = F(v["kpi_bg"])
        dash.cell(row=57, column=c).font = Fn(v["accent"], bold=True, size=9, name=fn)

    # Fill remaining cols with bg
    for r in range(44, 58):
        for c in range(6, ncols + 1):
            dash.cell(row=r, column=c).fill = F(v["bg"])

    # Column widths
    col_widths(dash, [3, 12, 4, 20, 4, 14, 4, 14, 4, 14, 4, 4, 4, 3])

    # ── Sheet: Budget ─────────────────────────────────────────────────────────
    bud = wb.create_sheet("Budget")
    bud.sheet_view.showGridLines = False
    bud.sheet_view.zoomScale = 90

    budget_cats = ["Housing", "Food & Groceries", "Transport",
                   "Utilities", "Entertainment", "Personal Care", "Health", "Other"]
    ncols_b = 3 + len(budget_cats) + 3  # year, month, income, cats, total, net, savings%
    _dash_header(bud, v, fn, "BUDGET DATA", "Enter your monthly income and expenses below", ncols_b)

    row_fill(bud, 5, 1, ncols_b, F(v["bg"]), height=10)

    headers_b = ["Year", "Month", "Income"] + budget_cats + ["Total Expenses", "Net Savings", "Savings Rate"]
    table_header(bud, 6, headers_b, v, fn)

    INC_COL  = 3   # C
    CAT_START = 4  # D
    CAT_END   = 4 + len(budget_cats) - 1  # K
    TOT_COL  = CAT_END + 1   # L
    NET_COL  = CAT_END + 2   # M
    SAV_COL  = CAT_END + 3   # N

    for mi, (mname, sample) in enumerate(zip(MONTHS, BUDGET_SAMPLE)):
        r = 7 + mi
        income = sample[0]
        cats   = sample[1:]
        total  = sum(cats)
        net    = income - total

        row_values = [year, mname, income] + list(cats) + [None, None, None]
        fmts_b = [None, None, num_fmt] + [num_fmt]*len(budget_cats) + [num_fmt, num_fmt, "0.0%"]

        table_row(bud, r, row_values, v, fn, col_start=1, alt=(mi%2==0), fmts=fmts_b)

        # Formulas for totals
        exp_range = f"{get_column_letter(CAT_START)}{r}:{get_column_letter(CAT_END)}{r}"
        bud.cell(row=r, column=TOT_COL).value  = f"=SUM({exp_range})"
        bud.cell(row=r, column=NET_COL).value  = f"=C{r}-{get_column_letter(TOT_COL)}{r}"
        bud.cell(row=r, column=SAV_COL).value  = f"=IF(C{r}=0,0,{get_column_letter(NET_COL)}{r}/C{r})"
        bud.cell(row=r, column=TOT_COL).number_format = num_fmt
        bud.cell(row=r, column=NET_COL).number_format = num_fmt
        bud.cell(row=r, column=SAV_COL).number_format = "0.0%"

    # Totals row
    tot_r = 19
    tot_vals = ["", "TOTALS"] + [f"=SUM({get_column_letter(c)}7:{get_column_letter(c)}18)"
                                  for c in range(INC_COL, SAV_COL)]
    fmts_tot = [None, None] + [num_fmt]*(len(headers_b)-3) + ["0.0%"]
    table_row(bud, tot_r, tot_vals + [""], v, fn, col_start=1, fmts=fmts_tot + [None])
    for c in range(1, len(headers_b) + 1):
        bud.cell(row=tot_r, column=c).fill = F(v["kpi_bg"])
        bud.cell(row=tot_r, column=c).font = Fn(v["accent"], bold=True, size=9, name=fn)
    bud.cell(row=tot_r, column=SAV_COL).value  = f"=AVERAGE(N7:N18)"
    bud.cell(row=tot_r, column=SAV_COL).number_format = "0.0%"

    # Input note
    row_fill(bud, 20, 1, ncols_b, F(v["bg"]), height=10)
    merge(bud, 21, 1, 21, ncols_b,
          "💡  Enter your figures in the yellow-highlighted cells. Totals, Net Savings and Savings Rate calculate automatically.",
          fill=F(v["kpi_bg"]),
          font=Fn(v["subtext"], italic=True, size=8, name=fn),
          align=Al("left", "center"))
    bud.row_dimensions[21].height = 18
    bud.cell(row=21, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Highlight input cells
    input_fill = PatternFill("solid", fgColor="FFFFF0")
    for r in range(7, 19):
        for c in range(INC_COL, CAT_END + 1):
            bud.cell(row=r, column=c).fill = input_fill

    col_widths(bud, [6, 13, 13] + [14]*8 + [14, 13, 13])
    bud.freeze_panes = "C7"

    # ── Now build charts referencing Budget sheet ─────────────────────────────
    # Bar chart: monthly income vs expenses
    inc_ref  = Reference(bud, min_col=INC_COL, min_row=6, max_row=18)
    exp_ref  = Reference(bud, min_col=TOT_COL, min_row=6, max_row=18)
    cats_ref = Reference(bud, min_col=2, min_row=7, max_row=18)
    chart_bar.add_data(inc_ref, titles_from_data=True)
    chart_bar.add_data(exp_ref, titles_from_data=True)
    chart_bar.set_categories(cats_ref)
    dash.add_chart(chart_bar, "A14")

    # Donut: category totals (row 19 = totals, cols D-K)
    cat_vals_ref   = Reference(bud, min_col=CAT_START, max_col=CAT_END, min_row=19)
    cat_labels_ref = Reference(bud, min_col=CAT_START, max_col=CAT_END, min_row=6)
    chart_donut.add_data(cat_vals_ref)
    chart_donut.set_categories(cat_labels_ref)
    dash.add_chart(chart_donut, "I14")

    return wb


# ─── Freelancer Template ──────────────────────────────────────────────────────

def build_freelancer_template(vibe_name, sym, num_fmt, year=None):
    v  = VIBES[vibe_name]
    fn = v["font"]
    if year is None: year = datetime.date.today().year
    ncols = 12
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Dashboard ─────────────────────────────────────────────────────────────
    dash = wb.create_sheet("Dashboard")
    dash.sheet_view.showGridLines = False
    dash.sheet_view.zoomScale = 90

    _dash_header(dash, v, fn,
                 f"FREELANCE FINANCE TRACKER  {year}",
                 "Income, expenses, profit & tax — all in one place",
                 ncols)

    cur_r = _currency_settings(dash, v, fn, 5, sym, ncols)

    row_fill(dash, 7, 1, ncols, F(v["bg"]), height=10)
    section_bar(dash, 8, 1, ncols, "  FINANCIAL SUMMARY", v, fn)

    dash.row_dimensions[9].height = 18
    dash.row_dimensions[10].height = 38
    dash.row_dimensions[11].height = 8

    kpis = [
        ("TOTAL REVENUE",   "=IFERROR(SUM('Income'!E2:E101),0)", 1, 3),
        ("TOTAL EXPENSES",  "=IFERROR(SUM('Expenses'!D2:D101),0)", 4, 6),
        ("NET PROFIT",      "=B10-E10", 7, 9),
        ("TAX ESTIMATE",    "=MAX(0,G10*0.2)", 10, 12),
    ]
    # Store KPI cell refs for Net Profit to reference
    kpi_cols = {}
    for i, (label, formula, cs, ce) in enumerate(kpis):
        kpi_box(dash, 9, cs, 11, ce, label, formula, v, fn, num_fmt=num_fmt)
        kpi_cols[label] = (cs, 10)  # (col, row)

    # Fix Net Profit formula to reference actual cells
    dash.cell(row=10, column=7).value = "=IFERROR(B10-E10,0)"
    dash.cell(row=10, column=10).value = "=IFERROR(MAX(0,H10*0.2),0)"

    row_fill(dash, 12, 1, ncols, F(v["bg"]), height=10)
    section_bar(dash, 13, 1, ncols, "  CHARTS & INSIGHTS", v, fn)
    for r in range(14, 40):
        row_fill(dash, r, 1, ncols, F(v["bg"]), height=15)

    row_fill(dash, 40, 1, ncols, F(v["bg"]), height=10)
    section_bar(dash, 41, 1, ncols, "  INCOME STATUS BREAKDOWN", v, fn)

    # Status KPIs: Paid vs Unpaid
    dash.row_dimensions[42].height = 18
    dash.row_dimensions[43].height = 36
    dash.row_dimensions[44].height = 8

    status_kpis = [
        ("PAID INVOICES",   "=IFERROR(COUNTIF('Income'!F2:F101,\"Paid\"),0)", 1, 4, "#,##0"),
        ("REVENUE RECEIVED",f"=IFERROR(SUMIF('Income'!F2:F101,\"Paid\",'Income'!E2:E101),0)", 5, 8, num_fmt),
        ("OUTSTANDING",     f"=IFERROR(SUMIF('Income'!F2:F101,\"Unpaid\",'Income'!E2:E101),0)", 9, 12, num_fmt),
    ]
    for label, formula, cs, ce, fmt in status_kpis:
        kpi_box(dash, 42, cs, 44, ce, label, formula, v, fn, num_fmt=fmt)

    col_widths(dash, [3, 14, 4, 14, 4, 14, 4, 14, 4, 14, 4, 3])

    # ── Income Sheet ─────────────────────────────────────────────────────────
    inc_ws = wb.create_sheet("Income")
    inc_ws.sheet_view.showGridLines = False
    inc_ws.sheet_view.zoomScale = 90

    _dash_header(inc_ws, v, fn, "INCOME LOG", "Record all client payments and invoices here", 6)
    row_fill(inc_ws, 5, 1, 6, F(v["bg"]), height=8)

    inc_headers = ["Date", "Client", "Project / Description", "Invoice #", "Amount", "Status"]
    table_header(inc_ws, 6, inc_headers, v, fn)

    dv_status = DataValidation(type="list", formula1='"Paid,Unpaid,Overdue"', showDropDown=False)
    inc_ws.add_data_validation(dv_status)

    # Conditional formatting: Paid=green, Overdue=red
    paid_fill    = PatternFill("solid", fgColor="D4EDDA" if vibe_name != "Dark & Minimal" else "1A3B2A")
    overdue_fill = PatternFill("solid", fgColor="F8D7DA" if vibe_name != "Dark & Minimal" else "3B1A1A")

    for ri, row_data in enumerate(FREELANCER_INCOME_SAMPLE):
        r = 7 + ri
        vals = list(row_data)
        fmts = [None, None, None, None, num_fmt if vals[3] else None, None]
        table_row(inc_ws, r, vals, v, fn, col_start=1, alt=(ri%2==0), fmts=fmts)
        if vals[3]:
            inc_ws.cell(row=r, column=1).number_format = "DD MMM YYYY"
        if vals[4]:
            dv_status.add(inc_ws.cell(row=r, column=6))
        inc_ws.conditional_formatting.add(f"A{r}:F{r}", FormulaRule(formula=[f'$F{r}="Paid"'],    fill=paid_fill))
        inc_ws.conditional_formatting.add(f"A{r}:F{r}", FormulaRule(formula=[f'$F{r}="Overdue"'], fill=overdue_fill))

    # Extend blank rows to 100
    for ri in range(len(FREELANCER_INCOME_SAMPLE), 94):
        r = 7 + ri
        table_row(inc_ws, r, ["","","","","",""], v, fn, col_start=1, alt=(ri%2==0),
                  fmts=[None, None, None, None, num_fmt, None])
        dv_status.add(inc_ws.cell(row=r, column=6))
        inc_ws.conditional_formatting.add(f"A{r}:F{r}", FormulaRule(formula=[f'$F{r}="Paid"'],    fill=paid_fill))
        inc_ws.conditional_formatting.add(f"A{r}:F{r}", FormulaRule(formula=[f'$F{r}="Overdue"'], fill=overdue_fill))

    col_widths(inc_ws, [14, 22, 30, 12, 14, 12])
    inc_ws.freeze_panes = "A7"

    # ── Expenses Sheet ────────────────────────────────────────────────────────
    exp_ws = wb.create_sheet("Expenses")
    exp_ws.sheet_view.showGridLines = False
    exp_ws.sheet_view.zoomScale = 90

    _dash_header(exp_ws, v, fn, "EXPENSE LOG", "Track all business expenses for tax & P&L", 5)
    row_fill(exp_ws, 5, 1, 5, F(v["bg"]), height=8)

    exp_cats_list = ["Software", "Equipment", "Travel", "Marketing", "Professional Services", "Other"]
    exp_headers = ["Date", "Category", "Description", "Amount"]
    table_header(exp_ws, 6, exp_headers, v, fn)

    dv_cat = DataValidation(type="list", formula1=f'"{",".join(exp_cats_list)}"', showDropDown=False)
    exp_ws.add_data_validation(dv_cat)

    for ri, row_data in enumerate(FREELANCER_EXPENSE_SAMPLE):
        r = 7 + ri
        vals = list(row_data)
        fmts = [None, None, None, num_fmt if vals[3] else None]
        table_row(exp_ws, r, vals, v, fn, col_start=1, alt=(ri%2==0), fmts=fmts)
        if vals[0]: exp_ws.cell(row=r, column=1).number_format = "DD MMM YYYY"
        dv_cat.add(exp_ws.cell(row=r, column=2))

    for ri in range(len(FREELANCER_EXPENSE_SAMPLE), 94):
        r = 7 + ri
        table_row(exp_ws, r, ["", "", "", ""], v, fn, alt=(ri%2==0),
                  fmts=[None, None, None, num_fmt])
        dv_cat.add(exp_ws.cell(row=r, column=2))

    col_widths(exp_ws, [14, 22, 32, 14, 4])
    exp_ws.freeze_panes = "A7"

    # ── Tax Summary Sheet ─────────────────────────────────────────────────────
    tax_ws = wb.create_sheet("Tax Summary")
    tax_ws.sheet_view.showGridLines = False
    tax_ws.sheet_view.zoomScale = 90

    _dash_header(tax_ws, v, fn, "TAX SUMMARY", "Estimated tax breakdown — always consult a professional", 6)
    row_fill(tax_ws, 5, 1, 6, F(v["bg"]), height=10)
    row_fill(tax_ws, 6, 1, 6, F(v["bg"]), height=8)

    tax_rows = [
        ("Total Revenue",         "=SUM('Income'!E2:E101)"),
        ("Total Allowable Expenses", "=SUM('Expenses'!D2:D101)"),
        ("Taxable Profit",        "=B8-B9"),
        ("",                      ""),
        ("Tax Rate",              0.20),
        ("Estimated Tax Owed",    "=B10*B12"),
        ("",                      ""),
        ("Take-Home (after tax)", "=B10-B13"),
    ]
    section_bar(tax_ws, 7, 1, 6, "  ESTIMATED TAX CALCULATION", v, fn)
    for ti, (label, val) in enumerate(tax_rows):
        r = 8 + ti
        is_total = label in ("Taxable Profit", "Estimated Tax Owed", "Take-Home (after tax)")
        bg = v["kpi_bg"] if is_total else (v["alt_row"] if ti % 2 == 0 else v["bg"])
        fc = v["accent"] if is_total else v["text"]
        tax_ws.row_dimensions[r].height = 20
        merge(tax_ws, r, 1, r, 3, label,
              fill=F(bg), font=Fn(fc, bold=is_total, size=9, name=fn),
              align=Al("left", "center"), border=Br(v["border"]))
        tax_ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c = tax_ws.cell(row=r, column=4)
        c.value = val
        c.fill  = F(bg)
        c.font  = Fn(fc, bold=is_total, size=9, name=fn)
        c.alignment = Al("center", "center")
        c.border = Br(v["border"])
        if isinstance(val, float) and val < 1: c.number_format = "0%"
        elif label: c.number_format = num_fmt

    merge(tax_ws, 17, 1, 17, 6,
          "⚠  This is an estimate only. Tax rates, allowances and reliefs vary. Always consult a qualified accountant.",
          fill=F(v["kpi_bg"]),
          font=Fn(v["subtext"], italic=True, size=8, name=fn),
          align=Al("left", "center"))
    tax_ws.row_dimensions[17].height = 18
    tax_ws.cell(row=17, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    col_widths(tax_ws, [3, 30, 10, 16, 3, 3])

    # ── Charts ────────────────────────────────────────────────────────────────
    # Donut: expense by category
    exp_chart = DoughnutChart()
    exp_chart.title = "Expenses by Category"
    exp_chart.style = v["chart_style"]
    exp_chart.holeSize = 55
    exp_chart.width = 14; exp_chart.height = 14
    exp_chart.legend.position = "r"
    # Build category totals helper in expenses sheet
    CAT_TOT_ROW = 105
    exp_ws.row_dimensions[CAT_TOT_ROW] = exp_ws.row_dimensions[CAT_TOT_ROW]
    for ci, cat in enumerate(exp_cats_list):
        exp_ws.cell(row=CAT_TOT_ROW, column=1 + ci, value=cat)
        exp_ws.cell(row=CAT_TOT_ROW + 1, column=1 + ci,
                    value=f'=SUMIF(B7:B100,"{cat}",D7:D100)')
        exp_ws.cell(row=CAT_TOT_ROW + 1, column=1 + ci).number_format = num_fmt

    cat_labels = Reference(exp_ws, min_col=1, max_col=len(exp_cats_list), min_row=CAT_TOT_ROW)
    cat_vals   = Reference(exp_ws, min_col=1, max_col=len(exp_cats_list), min_row=CAT_TOT_ROW + 1)
    exp_chart.add_data(cat_vals)
    exp_chart.set_categories(cat_labels)
    dash.add_chart(exp_chart, "A14")

    return wb


# ─── Savings Goals Template ───────────────────────────────────────────────────

def build_savings_template(vibe_name, sym, num_fmt, year=None):
    v  = VIBES[vibe_name]
    fn = v["font"]
    if year is None: year = datetime.date.today().year
    ncols = 12
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Dashboard ─────────────────────────────────────────────────────────────
    dash = wb.create_sheet("Dashboard")
    dash.sheet_view.showGridLines = False
    dash.sheet_view.zoomScale = 90

    _dash_header(dash, v, fn,
                 f"SAVINGS GOALS TRACKER  {year}",
                 "Visualise your progress and stay on track",
                 ncols)

    cur_r = _currency_settings(dash, v, fn, 5, sym, ncols)

    row_fill(dash, 7, 1, ncols, F(v["bg"]), height=10)
    section_bar(dash, 8, 1, ncols, "  OVERVIEW", v, fn)

    dash.row_dimensions[9].height = 18
    dash.row_dimensions[10].height = 38
    dash.row_dimensions[11].height = 8

    kpis_s = [
        ("TOTAL SAVED",     "=IFERROR(SUM('Goals'!C2:C20),0)", 1, 3),
        ("TOTAL TARGET",    "=IFERROR(SUM('Goals'!B2:B20),0)", 4, 6),
        ("STILL NEEDED",    "=IFERROR(SUM('Goals'!B2:B20)-SUM('Goals'!C2:C20),0)", 7, 9),
        ("ACTIVE GOALS",    "=IFERROR(COUNTA('Goals'!A2:A20),0)", 10, 12),
    ]
    for label, formula, cs, ce in kpis_s:
        fmt = "#,##0" if "GOALS" in label else num_fmt
        kpi_box(dash, 9, cs, 11, ce, label, formula, v, fn, num_fmt=fmt)

    row_fill(dash, 12, 1, ncols, F(v["bg"]), height=10)
    section_bar(dash, 13, 1, ncols, "  CHARTS & PROGRESS", v, fn)
    for r in range(14, 40):
        row_fill(dash, r, 1, ncols, F(v["bg"]), height=15)

    row_fill(dash, 40, 1, ncols, F(v["bg"]), height=10)

    # Goals summary table on dashboard
    section_bar(dash, 41, 1, ncols, "  GOALS AT A GLANCE", v, fn)
    tbl_h = ["Goal", "Target", "Saved", "Still Needed", "Monthly", "% Complete", "Est. Date"]
    table_header(dash, 42, tbl_h, v, fn, col_start=1)

    for gi in range(len(GOALS_SAMPLE)):
        r = 43 + gi
        gr = gi + 2  # Goals sheet row
        row_data = [
            f"='Goals'!A{gr}",
            f"='Goals'!B{gr}",
            f"='Goals'!C{gr}",
            f"=IF('Goals'!B{gr}>0,'Goals'!B{gr}-'Goals'!C{gr},\"\")",
            f"='Goals'!D{gr}",
            f"=IFERROR('Goals'!C{gr}/'Goals'!B{gr},0)",
            f"=IFERROR(TODAY()+('Goals'!B{gr}-'Goals'!C{gr})/'Goals'!D{gr}*30.44,\"\")",
        ]
        fmts = [None, num_fmt, num_fmt, num_fmt, num_fmt, "0.0%", "MMM YYYY"]
        table_row(dash, r, row_data, v, fn, col_start=1, alt=(gi%2==0), fmts=fmts)

    # Fill remaining cols with bg
    for r in range(42, 43 + len(GOALS_SAMPLE)):
        for c in range(8, ncols + 1):
            dash.cell(row=r, column=c).fill = F(v["bg"])

    col_widths(dash, [3, 22, 14, 12, 14, 12, 12, 4, 4, 4, 4, 3])

    # ── Goals Sheet ───────────────────────────────────────────────────────────
    goals_ws = wb.create_sheet("Goals")
    goals_ws.sheet_view.showGridLines = False
    goals_ws.sheet_view.zoomScale = 90
    ncols_g = 8

    _dash_header(goals_ws, v, fn, "SAVINGS GOALS", "Enter your goals and update your savings regularly", ncols_g)
    row_fill(goals_ws, 5, 1, ncols_g, F(v["bg"]), height=8)

    g_headers = ["Goal Name", "Target Amount", "Amount Saved", "Monthly Contribution",
                 "Months Remaining", "% Complete", "Est. Completion", "Notes"]
    table_header(goals_ws, 6, g_headers, v, fn)

    for gi, (name, target, current, monthly) in enumerate(GOALS_SAMPLE):
        r = 7 + gi
        fmts_g = [None, num_fmt, num_fmt, num_fmt, "#,##0", "0.0%", "MMM YYYY", None]
        row_data = [
            name, target, current, monthly,
            f"=IF(D{r}=0,\"-\",ROUNDUP((B{r}-C{r})/D{r},0))",
            f"=IFERROR(C{r}/B{r},0)",
            f"=IFERROR(TODAY()+(B{r}-C{r})/D{r}*30.44,\"\")",
            "",
        ]
        table_row(goals_ws, r, row_data, v, fn, col_start=1, alt=(gi%2==0), fmts=fmts_g)

    # Blank rows for more goals
    for gi in range(len(GOALS_SAMPLE), 19):
        r = 7 + gi
        fmts_g = [None, num_fmt, num_fmt, num_fmt, "#,##0", "0.0%", "MMM YYYY", None]
        row_data = ["", "", "", "",
                    f"=IF(D{r}=0,\"-\",ROUNDUP((B{r}-C{r})/D{r},0))",
                    f"=IFERROR(C{r}/B{r},0)",
                    f"=IFERROR(TODAY()+(B{r}-C{r})/D{r}*30.44,\"\")",
                    ""]
        table_row(goals_ws, r, row_data, v, fn, col_start=1, alt=(gi%2==0), fmts=fmts_g)

    # Progress data bars on % Complete column (col F = 6)
    db = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                     color=v["accent"])
    goals_ws.conditional_formatting.add(f"F7:F25", db)

    # Highlight input cells
    for r in range(7, 26):
        for c in [2, 3, 4]:
            goals_ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFFFF0")

    col_widths(goals_ws, [22, 16, 16, 20, 18, 14, 16, 20])
    goals_ws.freeze_panes = "A7"

    # ── Contribution Log Sheet ────────────────────────────────────────────────
    log_ws = wb.create_sheet("Contribution Log")
    log_ws.sheet_view.showGridLines = False
    log_ws.sheet_view.zoomScale = 90

    _dash_header(log_ws, v, fn, "CONTRIBUTION LOG", "Track every deposit toward your goals", 5)
    row_fill(log_ws, 5, 1, 5, F(v["bg"]), height=8)

    log_headers = ["Date", "Goal", "Amount Added", "Running Total", "Notes"]
    table_header(log_ws, 6, log_headers, v, fn)

    goal_names = [g[0] for g in GOALS_SAMPLE]
    dv_goal = DataValidation(type="list",
                             formula1=f'"{",".join(goal_names)}"',
                             showDropDown=False)
    log_ws.add_data_validation(dv_goal)

    sample_log = [
        ("01 Jan 2026", "Emergency Fund",   500, "January contribution"),
        ("01 Jan 2026", "Holiday — 2026",   350, "January contribution"),
        ("01 Feb 2026", "Emergency Fund",   500, "February contribution"),
        ("01 Feb 2026", "New Laptop",       150, "February contribution"),
        ("01 Mar 2026", "Emergency Fund",   500, "March contribution"),
        ("01 Mar 2026", "House Deposit",    800, "March contribution"),
    ]
    for li, row_data in enumerate(sample_log):
        r = 7 + li
        fmts_l = [None, None, num_fmt, num_fmt, None]
        table_row(log_ws, r, list(row_data), v, fn, alt=(li%2==0), fmts=fmts_l)
        log_ws.cell(row=r, column=1).number_format = "DD MMM YYYY"
        dv_goal.add(log_ws.cell(row=r, column=2))

    for li in range(len(sample_log), 94):
        r = 7 + li
        table_row(log_ws, r, ["", "", "", "", ""], v, fn, alt=(li%2==0),
                  fmts=[None, None, num_fmt, num_fmt, None])
        dv_goal.add(log_ws.cell(row=r, column=2))

    col_widths(log_ws, [14, 22, 16, 16, 28])
    log_ws.freeze_panes = "A7"

    # ── Charts ────────────────────────────────────────────────────────────────
    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.style = v["chart_style"]
    bar_chart.title = "Savings Progress"
    bar_chart.y_axis.title = "Goal"
    bar_chart.x_axis.title = "Amount"
    bar_chart.width = 20; bar_chart.height = 14
    bar_chart.legend.position = "b"

    g_names = Reference(goals_ws, min_col=1, min_row=7, max_row=7 + len(GOALS_SAMPLE) - 1)
    g_saved = Reference(goals_ws, min_col=3, min_row=6, max_row=7 + len(GOALS_SAMPLE) - 1)
    g_target= Reference(goals_ws, min_col=2, min_row=6, max_row=7 + len(GOALS_SAMPLE) - 1)
    bar_chart.add_data(g_saved,  titles_from_data=True)
    bar_chart.add_data(g_target, titles_from_data=True)
    bar_chart.set_categories(g_names)
    dash.add_chart(bar_chart, "A14")

    return wb


# ─── TUI ─────────────────────────────────────────────────────────────────────

def run_tui():
    console.clear()
    console.print(Panel(
        Text.assemble(
            ("FINANCE TEMPLATE GENERATOR\n", "bold cyan"),
            ("Create premium, sellable Excel spreadsheet templates", "dim"),
        ),
        padding=(1, 8), style="cyan", box=box.ROUNDED
    ))

    # Template type
    console.print()
    t_type = questionary.select(
        "What type of template are you generating?",
        choices=[f"{k}  —  {v}" for k, v in TEMPLATE_TYPES.items()]
    ).ask()
    template_key = t_type.split("  —  ")[0]

    # Vibe
    console.print()
    vibe_choice = questionary.select(
        "Choose a vibe / visual theme:",
        choices=[f"{k}  —  {v['desc']}" for k, v in VIBES.items()]
    ).ask()
    vibe_key = vibe_choice.split("  —  ")[0]

    # Currency
    console.print()
    cur_choice = questionary.select(
        "Default currency for this template?",
        choices=list(CURRENCIES.keys())
    ).ask()
    sym, num_fmt = CURRENCIES[cur_choice]

    # Year
    console.print()
    year_str = questionary.text(
        "Year this template covers:", default=str(datetime.date.today().year)
    ).ask()
    try: year = int(year_str)
    except: year = datetime.date.today().year

    # Filename
    slug = template_key.lower().replace(" / ", "-").replace(" ", "-")
    vibe_slug = vibe_key.lower().replace(" & ", "-").replace(" ", "-")
    default_name = f"{slug}_{vibe_slug}_{year}"
    console.print()
    fname = questionary.text("Output filename (no extension):", default=default_name).ask()
    if not fname: fname = default_name

    # Preview
    console.print()
    font = VIBES[vibe_key]["font"]
    console.print(Panel(
        f"[bold]Ready to generate[/bold]\n\n"
        f"  Template  : [cyan]{template_key}[/cyan]\n"
        f"  Vibe      : [cyan]{vibe_key}[/cyan]  (font: {font})\n"
        f"  Currency  : [cyan]{sym}[/cyan]\n"
        f"  Year      : [cyan]{year}[/cyan]\n"
        f"  File      : [cyan]{fname}.xlsx[/cyan]",
        style="green", padding=(0, 4), box=box.ROUNDED
    ))

    if not questionary.confirm("Generate now?", default=True).ask():
        console.print("[yellow]Cancelled.[/yellow]")
        sys.exit(0)

    return template_key, vibe_key, sym, num_fmt, year, fname


# ─── Answer log ───────────────────────────────────────────────────────────────

def log_answers(template_key, vibe_key, sym, year, fname):
    last = 0
    if LOG_FILE.exists():
        for line in reversed(LOG_FILE.read_text("utf-8").splitlines()):
            line = line.strip()
            if line and line[0].isdigit():
                try: last = int(line.split(".")[0]); break
                except: pass

    n = last
    entries = []
    def add(t): nonlocal n; n+=1; entries.append(f"{n}. {t}")

    add(f"[Finance Template v2] Template type: {template_key}")
    add(f"[Finance Template v2] Vibe: {vibe_key}  (font: {VIBES[vibe_key]['font']})")
    add(f"[Finance Template v2] Currency: {sym}")
    add(f"[Finance Template v2] Year: {year}")
    add(f"[Finance Template v2] Output file: {fname}.xlsx")

    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write("\n".join(entries) + "\n")


# ─── Main ─────────────────────────────────────────────────────────────────────

BUILDERS = {
    "Personal Budget Tracker": build_budget_template,
    "Freelancer / Self-Employed": build_freelancer_template,
    "Savings Goals Planner": build_savings_template,
}

def main():
    template_key, vibe_key, sym, num_fmt, year, fname = run_tui()

    console.print()
    console.print("[dim]Building workbook...[/dim]")

    wb = BUILDERS[template_key](vibe_key, sym, num_fmt, year)

    OUTPUT_DIR.mkdir(exist_ok=True)
    out_path = OUTPUT_DIR / f"{fname}.xlsx"
    wb.save(str(out_path))

    log_answers(template_key, vibe_key, sym, year, fname)

    console.print()
    console.print(Panel(
        f"[bold green]Done![/bold green]\n\n"
        f"  [cyan]{out_path}[/cyan]\n\n"
        f"  [dim]Light sample data is pre-filled so charts render immediately.\n"
        f"  Customers delete the sample rows and enter their own data.[/dim]",
        style="green", padding=(0, 4), box=box.ROUNDED
    ))

    if questionary.confirm("Open the file now?", default=True).ask():
        os.startfile(str(out_path))


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n[yellow]Aborted.[/yellow]")
        sys.exit(0)
