"""
Excel Finance Spreadsheet Generator
Run: python excel_finance_app-v1.py
Requires: pip install openpyxl rich questionary pywin32
"""

import os
import sys
import datetime
import pathlib

import questionary
from rich.console import Console
from rich.panel import Panel
from rich.text import Text
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

console = Console()

# ── Constants ────────────────────────────────────────────────────────────────

CURRENCIES = {
    "GBP £  (British Pound)":      ("GBP", "£",   '£#,##0.00'),
    "EUR €  (Euro)":               ("EUR", "€",   '€#,##0.00'),
    "USD $  (US Dollar)":          ("USD", "$",   '$#,##0.00'),
    "HUF Ft (Hungarian Forint)":   ("HUF", "Ft",  '#,##0\\ "Ft"'),
    "JPY ¥  (Japanese Yen)":       ("JPY", "¥",   '¥#,##0'),
    "CAD C$ (Canadian Dollar)":    ("CAD", "C$",  '"C$"#,##0.00'),
    "AUD A$ (Australian Dollar)":  ("AUD", "A$",  '"A$"#,##0.00'),
    "CHF    (Swiss Franc)":        ("CHF", "CHF", '"CHF"\\ #,##0.00'),
    "PLN zł (Polish Zloty)":       ("PLN", "zł",  '#,##0.00\\ "zł"'),
    "CZK Kč (Czech Koruna)":       ("CZK", "Kč",  '#,##0\\ "Kč"'),
}

THEMES = {
    "Midnight  — dark navy + cyan":     {
        "bg": "1B2838", "header_bg": "0D1B2A", "alt_row": "243447",
        "accent": "00B4D8", "totals_bg": "162232",
        "font": "FFFFFF", "header_font": "FFFFFF", "kpi_title": "90CAF9",
    },
    "Minimal   — white + blue":          {
        "bg": "FFFFFF", "header_bg": "E5E7EB", "alt_row": "F3F4F6",
        "accent": "2563EB", "totals_bg": "DBEAFE",
        "font": "111827", "header_font": "1E3A5F", "kpi_title": "374151",
    },
    "Forest    — dark green + sage":     {
        "bg": "1B4332", "header_bg": "081C15", "alt_row": "2D6A4F",
        "accent": "52B788", "totals_bg": "1B4332",
        "font": "D8F3DC", "header_font": "FFFFFF", "kpi_title": "95D5B2",
    },
    "Crimson   — dark red + coral":      {
        "bg": "1A0A0A", "header_bg": "3B0000", "alt_row": "2D1010",
        "accent": "EF4444", "totals_bg": "3B0000",
        "font": "FECACA", "header_font": "FFFFFF", "kpi_title": "FCA5A5",
    },
}

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

BUDGET_CATS = [
    "Rent / Mortgage", "Food & Groceries", "Transport",
    "Utilities", "Subscriptions", "Entertainment", "Health", "Other",
]

BIZ_EXPENSE_CATS = [
    "Equipment", "Software", "Travel", "Marketing",
    "Professional Services", "Other",
]

OUTPUT_DIR = pathlib.Path(__file__).parent / "output"
LOG_FILE   = pathlib.Path(__file__).parent / "answers_log.txt"

# ── TUI helpers ──────────────────────────────────────────────────────────────

def header(title, sub=""):
    body = f"[bold white]{title}[/bold white]"
    if sub:
        body += f"\n[dim]{sub}[/dim]"
    console.print()
    console.print(Panel(body, style="cyan", padding=(0, 6)))
    console.print()

def ask(prompt, default=""):
    return questionary.text(prompt, default=str(default)).ask() or str(default)

def sel(prompt, choices):
    return questionary.select(prompt, choices=choices).ask()

def chk(prompt, choices):
    result = questionary.checkbox(prompt, choices=choices).ask()
    return result or []

def ok(prompt, default=True):
    return questionary.confirm(prompt, default=default).ask()

# ── Input collection ─────────────────────────────────────────────────────────

def collect_setup():
    header("EXCEL FINANCE GENERATOR", "Step 1 of 6 — Setup")
    today = datetime.date.today().isoformat()

    fname = ask("Output filename (no extension):", default=f"finances_{today}")

    cur_key = sel("Currency:", list(CURRENCIES.keys()))
    cur = CURRENCIES[cur_key]

    th_key = sel("Colour theme:", list(THEMES.keys()))
    theme = THEMES[th_key]

    layout = sel("Workbook layout:", [
        "Dashboard + separate sheets  (recommended)",
        "All-in-one single sheet",
    ])
    multi = "Dashboard" in layout

    return dict(fname=fname, cur_key=cur_key, cur=cur,
                th_key=th_key, theme=theme, multi=multi)


def collect_modules():
    header("MODULES", "Step 2 of 6 — Select sections")
    mods = chk("Include modules:", [
        "Personal Budget",
        "Investments",
        "Savings Goals",
        "Business / Freelance",
    ])
    if not mods:
        console.print("[red]Select at least one module.[/red]")
        return collect_modules()
    return mods


def collect_budget(sym):
    header("PERSONAL BUDGET", "Step 3 — Weekly income & expenses")
    cats = chk("Expense categories:", BUDGET_CATS)
    if not cats:
        cats = ["Other"]
    extra = ask("Custom categories (comma-separated, blank = none):")
    if extra.strip():
        cats += [c.strip() for c in extra.split(",") if c.strip()]
    income = float(ask(f"Your weekly income ({sym}):", default="0") or 0)
    return dict(categories=cats, weekly_income=income)


def collect_investments(sym):
    header("INVESTMENTS", "Step 4 — Stocks & savings accounts")
    stocks, savings = [], []
    if ok("Add stock / ETF holdings?"):
        while True:
            ticker  = ask("Ticker symbol (e.g. VWRL):").strip().upper()
            buy     = float(ask(f"Buy price per share ({sym}):", default="0") or 0)
            qty     = float(ask("Quantity held:", default="0") or 0)
            cur_p   = float(ask(f"Current price per share ({sym}):", default="0") or 0)
            stocks.append(dict(ticker=ticker, buy=buy, qty=qty, cur=cur_p))
            if not ok("Add another holding?", default=False):
                break
    if ok("Add savings accounts / ISAs / bonds?"):
        while True:
            name    = ask("Account name (e.g. Cash ISA):").strip()
            balance = float(ask(f"Current balance ({sym}):", default="0") or 0)
            rate    = float(ask("Annual interest rate (%):", default="0") or 0)
            savings.append(dict(name=name, balance=balance, rate=rate))
            if not ok("Add another account?", default=False):
                break
    return dict(stocks=stocks, savings=savings)


def collect_goals(sym):
    header("SAVINGS GOALS", "Step 5 — Track financial targets")
    goals = []
    while True:
        name    = ask("Goal name (e.g. Emergency Fund):").strip()
        target  = float(ask(f"Target amount ({sym}):", default="0") or 0)
        current = float(ask(f"Amount saved so far ({sym}):", default="0") or 0)
        monthly = float(ask(f"Monthly contribution ({sym}):", default="0") or 0)
        goals.append(dict(name=name, target=target, current=current, monthly=monthly))
        if not ok("Add another goal?", default=False):
            break
    return dict(goals=goals)


def collect_business(sym):
    header("BUSINESS / FREELANCE", "Step 6 — Income & expense tracking")
    rate_type = sel("How do you charge?", ["Day rate", "Hourly rate", "Project-based"])
    rate      = float(ask(f"Rate ({sym}):", default="0") or 0)
    tax_rate  = float(ask("Estimated tax rate (%):", default="20") or 20)
    exp_cats  = chk("Expense categories:", BIZ_EXPENSE_CATS)
    if not exp_cats:
        exp_cats = ["Other"]
    return dict(rate_type=rate_type, rate=rate, tax_rate=tax_rate, exp_cats=exp_cats)


# ── Answer log ───────────────────────────────────────────────────────────────

def log_answers(setup, mods, budget, investments, goals, business):
    last = 0
    if LOG_FILE.exists():
        for line in reversed(LOG_FILE.read_text("utf-8").splitlines()):
            line = line.strip()
            if line and line[0].isdigit():
                try:
                    last = int(line.split(".")[0]); break
                except ValueError:
                    pass

    entries, n = [], last
    def add(t):
        nonlocal n; n += 1; entries.append(f"{n}. {t}")

    sym = setup["cur"][1]
    add(f"Setup — Filename: {setup['fname']}")
    add(f"Setup — Currency: {setup['cur_key']}")
    add(f"Setup — Theme: {setup['th_key']}")
    add(f"Setup — Layout: {'Dashboard + sheets' if setup['multi'] else 'All-in-one'}")
    add(f"Modules: {', '.join(mods)}")

    if budget:
        add(f"Budget — Weekly income: {sym}{budget['weekly_income']}")
        add(f"Budget — Categories: {', '.join(budget['categories'])}")

    if investments:
        add(f"Investments — Stocks: {len(investments['stocks'])} holding(s)")
        add(f"Investments — Savings accounts: {len(investments['savings'])} account(s)")

    if goals:
        names = ", ".join(g["name"] for g in goals["goals"])
        add(f"Savings Goals — {len(goals['goals'])} goal(s): {names}")

    if business:
        add(f"Business — Rate: {sym}{business['rate']} ({business['rate_type']}), Tax: {business['tax_rate']}%")

    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write("\n".join(entries) + "\n")

# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex6):
    return PatternFill("solid", fgColor=hex6)

def _font(hex6, bold=False, size=11, italic=False):
    return Font(color=hex6, bold=bold, size=size, italic=italic)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(color="AAAAAA"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def style_row(ws, row, theme, ncols, start=1, kind="data", alt=False):
    if kind == "header":
        f, bg = _font(theme["header_font"], bold=True, size=10), theme["header_bg"]
    elif kind == "total":
        f, bg = _font(theme["accent"], bold=True, size=10), theme["totals_bg"]
    else:
        f = _font(theme["font"], size=10)
        bg = theme["alt_row"] if alt else theme["bg"]
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(bg)
        cell.font = f
        cell.alignment = _align(wrap=True)
        cell.border = _border()

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def merge_title(ws, text, row, start_col, end_col, theme):
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col)
    cell.value = text
    cell.fill = _fill(theme["header_bg"])
    cell.font = _font(theme["accent"], bold=True, size=14)
    cell.alignment = _align()
    ws.row_dimensions[row].height = 28

def money_fmt(ws, row, col, fmt):
    ws.cell(row=row, column=col).number_format = fmt

def apply_money_fmt_range(ws, min_row, max_row, cols, fmt):
    for r in range(min_row, max_row + 1):
        for c in cols:
            ws.cell(row=r, column=c).number_format = fmt

# ── Budget sheet ─────────────────────────────────────────────────────────────

def build_budget(ws, budget, theme, cfmt):
    cats = budget["categories"]
    ncols = 3 + len(cats) + 2  # Week#, Date, Income, [cats], TotalExp, Net

    # Title
    merge_title(ws, "Personal Budget — Weekly Tracker", 1, 1, ncols, theme)

    # Headers row 3
    headers = ["Week", "Week Start", "Income"] + cats + ["Total Expenses", "Net"]
    style_row(ws, 3, theme, ncols, kind="header")
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)

    # 52 weekly rows (4-55), seed current week income
    today = datetime.date.today()
    year_start = datetime.date(today.year, 1, 4)  # first Monday on/after Jan 1
    while year_start.weekday() != 0:
        year_start += datetime.timedelta(days=1)
    cur_week = today.isocalendar()[1]

    exp_start_col = 4  # column index where categories start
    exp_end_col   = 3 + len(cats)
    tot_col       = 3 + len(cats) + 1
    net_col       = 3 + len(cats) + 2

    for w in range(1, 53):
        row = w + 3
        week_date = year_start + datetime.timedelta(weeks=w - 1)
        style_row(ws, row, theme, ncols, alt=(w % 2 == 0))
        ws.cell(row=row, column=1, value=w)
        ws.cell(row=row, column=2, value=week_date).number_format = "DD MMM YYYY"
        if w == cur_week:
            ws.cell(row=row, column=3, value=budget["weekly_income"])
        else:
            ws.cell(row=row, column=3, value=0)
        for c in range(exp_start_col, exp_end_col + 1):
            ws.cell(row=row, column=c, value=0)
        # Total expenses formula
        exp_range = f"{get_column_letter(exp_start_col)}{row}:{get_column_letter(exp_end_col)}{row}"
        ws.cell(row=row, column=tot_col, value=f"=SUM({exp_range})")
        # Net formula
        ws.cell(row=row, column=net_col,
                value=f"={get_column_letter(3)}{row}-{get_column_letter(tot_col)}{row}")

    # Apply currency format to income/expense/total/net columns
    money_cols = [3] + list(range(exp_start_col, exp_end_col + 1)) + [tot_col, net_col]
    apply_money_fmt_range(ws, 4, 55, money_cols, cfmt)
    ws.cell(row=3, column=2).number_format = "DD MMM YYYY"

    # Monthly rollup — row 57 header, 58-69 months, 70 totals
    ROW_MHEAD = 57
    ROW_MDATA = 58
    merge_title(ws, "Monthly Rollup", ROW_MHEAD - 1, 1, 4, theme)
    style_row(ws, ROW_MHEAD, theme, 4, kind="header")
    for i, h in enumerate(["Month", "Income", "Total Expenses", "Net"], 1):
        ws.cell(row=ROW_MHEAD, column=i, value=h)

    date_col_letter = "B"
    inc_col_letter  = "C"
    exp_col_letter  = get_column_letter(tot_col)
    net_col_letter  = get_column_letter(net_col)

    for mi, month in enumerate(MONTHS):
        row = ROW_MDATA + mi
        style_row(ws, row, theme, 4, alt=(mi % 2 == 0))
        ws.cell(row=row, column=1, value=month)
        # SUMIFS: sum Income where week start date falls in this month
        month_num = mi + 1
        inc_formula = (
            f'=SUMPRODUCT((MONTH({date_col_letter}4:{date_col_letter}55)={month_num})'
            f'*({inc_col_letter}4:{inc_col_letter}55))'
        )
        exp_formula = (
            f'=SUMPRODUCT((MONTH({date_col_letter}4:{date_col_letter}55)={month_num})'
            f'*({exp_col_letter}4:{exp_col_letter}55))'
        )
        net_formula = (
            f'=SUMPRODUCT((MONTH({date_col_letter}4:{date_col_letter}55)={month_num})'
            f'*({net_col_letter}4:{net_col_letter}55))'
        )
        ws.cell(row=row, column=2, value=inc_formula)
        ws.cell(row=row, column=3, value=exp_formula)
        ws.cell(row=row, column=4, value=net_formula)

    apply_money_fmt_range(ws, ROW_MDATA, ROW_MDATA + 11, [2, 3, 4], cfmt)

    # Total row
    style_row(ws, 70, theme, 4, kind="total")
    ws.cell(row=70, column=1, value="TOTAL")
    for c, col in enumerate([2, 3, 4], 2):
        ws.cell(row=70, column=col,
                value=f"=SUM({get_column_letter(col)}{ROW_MDATA}:{get_column_letter(col)}{ROW_MDATA+11})")
        ws.cell(row=70, column=col).number_format = cfmt

    # Bar chart: Income vs Expenses by month
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Income vs Expenses by Month"
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Month"
    chart.width = 20
    chart.height = 12

    cats_ref = Reference(ws, min_col=1, min_row=ROW_MDATA, max_row=ROW_MDATA + 11)
    inc_ref  = Reference(ws, min_col=2, min_row=ROW_MHEAD, max_row=ROW_MDATA + 11)
    exp_ref  = Reference(ws, min_col=3, min_row=ROW_MHEAD, max_row=ROW_MDATA + 11)
    chart.add_data(inc_ref, titles_from_data=True)
    chart.add_data(exp_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"F{ROW_MHEAD}")

    # Column widths
    widths = [7, 14, 14] + [14] * len(cats) + [16, 12]
    set_col_widths(ws, widths)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return ROW_MHEAD, ROW_MDATA  # expose for dashboard reference

# ── Investments sheet ─────────────────────────────────────────────────────────

def build_investments(ws, investments, theme, cfmt):
    stocks  = investments["stocks"]
    savings = investments["savings"]

    merge_title(ws, "Investments", 1, 1, 7, theme)

    # Stocks section
    ROW_SHEAD = 3
    style_row(ws, ROW_SHEAD, theme, 7, kind="header")
    for i, h in enumerate(
        ["Ticker", "Buy Price", "Quantity", "Current Price", "Total Value", "P&L", "% Change"], 1
    ):
        ws.cell(row=ROW_SHEAD, column=i, value=h)

    ROW_SDATA = ROW_SHEAD + 1
    for si, s in enumerate(stocks if stocks else []):
        row = ROW_SDATA + si
        style_row(ws, row, theme, 7, alt=(si % 2 == 0))
        ws.cell(row=row, column=1, value=s["ticker"])
        ws.cell(row=row, column=2, value=s["buy"]).number_format = cfmt
        ws.cell(row=row, column=3, value=s["qty"])
        ws.cell(row=row, column=4, value=s["cur"]).number_format = cfmt
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}").number_format = cfmt
        ws.cell(row=row, column=6, value=f"=E{row}-(B{row}*C{row})").number_format = cfmt
        ws.cell(row=row, column=7, value=f"=IF(B{row}*C{row}=0,0,(E{row}-(B{row}*C{row}))/(B{row}*C{row}))").number_format = "0.00%"

    stock_end = ROW_SDATA + max(len(stocks), 1) - 1

    # Conditional formatting: green P&L if > 0, red if < 0
    if stocks:
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        green_fill = PatternFill("solid", fgColor="1A4731")
        red_fill   = PatternFill("solid", fgColor="4B1313")
        pl_range = f"F{ROW_SDATA}:F{stock_end}"
        ws.conditional_formatting.add(pl_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill))
        ws.conditional_formatting.add(pl_range,
            CellIsRule(operator="lessThan",    formula=["0"], fill=red_fill))

    # Total row for stocks
    tot_row = stock_end + 1
    style_row(ws, tot_row, theme, 7, kind="total")
    ws.cell(row=tot_row, column=1, value="TOTAL")
    for c in [5, 6]:
        ws.cell(row=tot_row, column=c,
                value=f"=SUM({get_column_letter(c)}{ROW_SDATA}:{get_column_letter(c)}{stock_end})")
        ws.cell(row=tot_row, column=c).number_format = cfmt

    # Portfolio allocation pie chart
    if stocks:
        chart = PieChart()
        chart.title = "Portfolio Allocation"
        chart.width = 16; chart.height = 12
        labels = Reference(ws, min_col=1, min_row=ROW_SDATA, max_row=stock_end)
        data   = Reference(ws, min_col=5, min_row=ROW_SDATA - 1, max_row=stock_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, f"I{ROW_SHEAD}")

    # Savings accounts section
    ROW_SAHEAD = tot_row + 2
    merge_title(ws, "Savings Accounts / ISAs / Bonds", ROW_SAHEAD - 1, 1, 4, theme)
    style_row(ws, ROW_SAHEAD, theme, 4, kind="header")
    for i, h in enumerate(["Account", "Balance", "Interest Rate %", "Annual Interest"], 1):
        ws.cell(row=ROW_SAHEAD, column=i, value=h)

    ROW_SADATA = ROW_SAHEAD + 1
    for ai, a in enumerate(savings if savings else []):
        row = ROW_SADATA + ai
        style_row(ws, row, theme, 4, alt=(ai % 2 == 0))
        ws.cell(row=row, column=1, value=a["name"])
        ws.cell(row=row, column=2, value=a["balance"]).number_format = cfmt
        ws.cell(row=row, column=3, value=a["rate"] / 100).number_format = "0.00%"
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}").number_format = cfmt

    sav_end = ROW_SADATA + max(len(savings), 1) - 1
    tot2 = sav_end + 1
    style_row(ws, tot2, theme, 4, kind="total")
    ws.cell(row=tot2, column=1, value="TOTAL")
    ws.cell(row=tot2, column=2, value=f"=SUM(B{ROW_SADATA}:B{sav_end})").number_format = cfmt
    ws.cell(row=tot2, column=4, value=f"=SUM(D{ROW_SADATA}:D{sav_end})").number_format = cfmt

    set_col_widths(ws, [16, 14, 16, 16, 14, 12, 10])
    ws.freeze_panes = f"A{ROW_SHEAD + 1}"
    ws.sheet_view.showGridLines = False
    return ROW_SDATA, stock_end, tot_row  # for dashboard

# ── Savings Goals sheet ───────────────────────────────────────────────────────

def build_savings(ws, goals_data, theme, cfmt):
    goals = goals_data["goals"]
    ncols = 7

    merge_title(ws, "Savings Goals", 1, 1, ncols, theme)

    style_row(ws, 3, theme, ncols, kind="header")
    for i, h in enumerate(
        ["Goal", "Target", "Saved", "Monthly Contribution",
         "Months to Goal", "% Complete", "Progress"], 1
    ):
        ws.cell(row=3, column=i, value=h)

    for gi, g in enumerate(goals):
        row = 4 + gi
        style_row(ws, row, theme, ncols, alt=(gi % 2 == 0))
        ws.cell(row=row, column=1, value=g["name"])
        ws.cell(row=row, column=2, value=g["target"]).number_format = cfmt
        ws.cell(row=row, column=3, value=g["current"]).number_format = cfmt
        ws.cell(row=row, column=4, value=g["monthly"]).number_format = cfmt
        ws.cell(row=row, column=5,
                value=f"=IF(D{row}=0,\"—\",ROUNDUP((B{row}-C{row})/D{row},0))")
        ws.cell(row=row, column=6,
                value=f"=IF(B{row}=0,0,C{row}/B{row})").number_format = "0.00%"
        ws.cell(row=row, column=7, value="")  # data bar placeholder

    # Data bar conditional formatting on "% Complete" column F
    data_end = 3 + len(goals)
    from openpyxl.formatting.rule import DataBarRule
    db_rule = DataBarRule(start_type="num", start_value=0,
                          end_type="num",   end_value=1,
                          color=theme["accent"])
    ws.conditional_formatting.add(f"F4:F{data_end}", db_rule)

    # Horizontal bar chart
    if goals:
        chart = BarChart()
        chart.type = "bar"  # horizontal
        chart.title = "Savings Goals Progress"
        chart.y_axis.title = "Goal"
        chart.x_axis.title = "Amount"
        chart.width = 20; chart.height = 12

        labels  = Reference(ws, min_col=1, min_row=4, max_row=data_end)
        current = Reference(ws, min_col=3, min_row=3, max_row=data_end)
        target  = Reference(ws, min_col=2, min_row=3, max_row=data_end)
        chart.add_data(current, titles_from_data=True)
        chart.add_data(target,  titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, "I3")

    set_col_widths(ws, [22, 14, 14, 20, 16, 14, 20])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ── Business sheet ────────────────────────────────────────────────────────────

def build_business(ws, biz, theme, cfmt):
    exp_cats = biz["exp_cats"]

    # ── Invoice tracker ──
    merge_title(ws, "Business / Freelance — Invoice Tracker", 1, 1, 6, theme)
    style_row(ws, 3, theme, 6, kind="header")
    for i, h in enumerate(["Client", "Date", "Invoice #", "Service", "Amount", "Status"], 1):
        ws.cell(row=3, column=i, value=h)

    INV_DATA = 4
    INV_END  = 33  # 30 invoice rows
    for r in range(INV_DATA, INV_END + 1):
        style_row(ws, r, theme, 6, alt=((r - INV_DATA) % 2 == 0))
        ws.cell(row=r, column=2).number_format = "DD MMM YYYY"
        ws.cell(row=r, column=5).number_format = cfmt

    # Status dropdown
    dv_status = DataValidation(type="list",
                               formula1='"Paid,Unpaid,Overdue"',
                               showDropDown=False)
    ws.add_data_validation(dv_status)
    for r in range(INV_DATA, INV_END + 1):
        dv_status.add(ws.cell(row=r, column=6))

    # Conditional formatting: Paid=green row, Overdue=red row
    green_fill = PatternFill("solid", fgColor="1A4731")
    red_fill   = PatternFill("solid", fgColor="4B1313")
    for r in range(INV_DATA, INV_END + 1):
        ws.conditional_formatting.add(
            f"A{r}:F{r}",
            FormulaRule(formula=[f'$F{r}="Paid"'],    fill=green_fill))
        ws.conditional_formatting.add(
            f"A{r}:F{r}",
            FormulaRule(formula=[f'$F{r}="Overdue"'], fill=red_fill))

    # Invoice total row
    TOT_INV = INV_END + 1
    style_row(ws, TOT_INV, theme, 6, kind="total")
    ws.cell(row=TOT_INV, column=1, value="TOTAL REVENUE")
    ws.cell(row=TOT_INV, column=5,
            value=f'=SUMIF(F{INV_DATA}:F{INV_END},"Paid",E{INV_DATA}:E{INV_END})')
    ws.cell(row=TOT_INV, column=5).number_format = cfmt

    # ── Expense log ──
    EXP_HEAD = TOT_INV + 2
    merge_title(ws, "Expenses", EXP_HEAD - 1, 1, 4, theme)
    style_row(ws, EXP_HEAD, theme, 4, kind="header")
    for i, h in enumerate(["Date", "Category", "Description", "Amount"], 1):
        ws.cell(row=EXP_HEAD, column=i, value=h)

    EXP_DATA = EXP_HEAD + 1
    EXP_END  = EXP_DATA + 29  # 30 expense rows
    for r in range(EXP_DATA, EXP_END + 1):
        style_row(ws, r, theme, 4, alt=((r - EXP_DATA) % 2 == 0))
        ws.cell(row=r, column=1).number_format = "DD MMM YYYY"
        ws.cell(row=r, column=4).number_format = cfmt

    # Category dropdown
    cat_list = ",".join(exp_cats)
    dv_cat = DataValidation(type="list",
                            formula1=f'"{cat_list}"',
                            showDropDown=False)
    ws.add_data_validation(dv_cat)
    for r in range(EXP_DATA, EXP_END + 1):
        dv_cat.add(ws.cell(row=r, column=2))

    TOT_EXP = EXP_END + 1
    style_row(ws, TOT_EXP, theme, 4, kind="total")
    ws.cell(row=TOT_EXP, column=1, value="TOTAL EXPENSES")
    ws.cell(row=TOT_EXP, column=4,
            value=f"=SUM(D{EXP_DATA}:D{EXP_END})").number_format = cfmt

    # ── P&L Summary ──
    PL_HEAD = TOT_EXP + 2
    merge_title(ws, "Profit & Loss Summary", PL_HEAD - 1, 1, 2, theme)

    tax = biz["tax_rate"] / 100
    pl_rows = [
        ("Total Revenue",   f"=E{TOT_INV}"),
        ("Total Expenses",  f"=D{TOT_EXP}"),
        ("Gross Profit",    f"=B{PL_HEAD+2}-B{PL_HEAD+3}"),
        (f"Tax ({biz['tax_rate']}%)", f"=B{PL_HEAD+4}*{tax}"),
        ("Take-Home Pay",   f"=B{PL_HEAD+4}-B{PL_HEAD+5}"),
    ]
    for i, (label, formula) in enumerate(pl_rows):
        row = PL_HEAD + 1 + i
        kind = "total" if i in (2, 4) else "data"
        style_row(ws, row, theme, 2, kind=kind, alt=(i % 2 == 0))
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula).number_format = cfmt

    # Monthly P&L line chart (using invoice totals by month — simplified)
    chart = LineChart()
    chart.title = "Monthly Invoice Revenue"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Month"
    chart.width = 20; chart.height = 12
    ws.add_chart(chart, "H3")  # placeholder — no data seeded

    set_col_widths(ws, [18, 14, 12, 22, 14, 10])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    return PL_HEAD

# ── Dashboard sheet ───────────────────────────────────────────────────────────

def build_dashboard(ws, setup, mods, budget_ws, inv_ws, goals_ws, biz_ws,
                    bud_rows, theme, cfmt):
    sym  = setup["cur"][1]
    ncols = 10

    # Title
    ws.row_dimensions[1].height = 36
    merge_title(ws, f"Financial Dashboard  |  {setup['fname']}  |  {datetime.date.today():%d %b %Y}",
                1, 1, ncols, theme)

    # Filter row (row 3)
    ws.row_dimensions[3].height = 24
    ws.merge_cells("A3:B3")
    ws.cell(row=3, column=1, value="Filter by month:").font = _font(theme["kpi_title"], bold=True)
    ws.cell(row=3, column=1).fill = _fill(theme["bg"])
    ws.cell(row=3, column=1).alignment = _align("right")

    ws.cell(row=3, column=3, value="All").fill = _fill(theme["header_bg"])
    ws.cell(row=3, column=3).font = _font(theme["accent"], bold=True)
    ws.cell(row=3, column=3).alignment = _align()
    ws.cell(row=3, column=3).border = _border(theme["accent"])

    dv = DataValidation(type="list",
                        formula1='"All,Jan,Feb,Mar,Apr,May,Jun,Jul,Aug,Sep,Oct,Nov,Dec"',
                        showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws["C3"])
    ws.defined_names["FilterPeriod"] = DefinedName("FilterPeriod", attr_text="Dashboard!$C$3")

    # KPI boxes — row 5 labels, row 6 values
    kpi_labels = ["Total Income", "Total Expenses", "Net Savings", "Portfolio Value"]
    kpi_cols   = [1, 3, 5, 7]
    for label, col in zip(kpi_labels, kpi_cols):
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        c_label = ws.cell(row=5, column=col, value=label)
        c_label.fill = _fill(theme["header_bg"])
        c_label.font = _font(theme["kpi_title"], bold=True, size=10)
        c_label.alignment = _align()
        c_val = ws.cell(row=6, column=col)
        c_val.fill = _fill(theme["bg"])
        c_val.font = _font(theme["accent"], bold=True, size=16)
        c_val.alignment = _align()
        c_val.number_format = cfmt

    # KPI formulas — pull from other sheets
    if bud_rows and budget_ws:
        mhead, mdata = bud_rows
        # Income: sum of monthly income col B (col 2)
        ws.cell(row=6, column=1).value = (
            f"=IFERROR(IF(C3=\"All\","
            f"SUM(Budget!B{mdata}:Budget!B{mdata+11}),"
            f"SUMIF(Budget!A{mdata}:Budget!A{mdata+11},C3,Budget!B{mdata}:Budget!B{mdata+11})),0)"
        )
        # Expenses: col 3
        ws.cell(row=6, column=3).value = (
            f"=IFERROR(IF(C3=\"All\","
            f"SUM(Budget!C{mdata}:Budget!C{mdata+11}),"
            f"SUMIF(Budget!A{mdata}:Budget!A{mdata+11},C3,Budget!C{mdata}:Budget!C{mdata+11})),0)"
        )
        # Net: col 4
        ws.cell(row=6, column=5).value = (
            f"=IFERROR(IF(C3=\"All\","
            f"SUM(Budget!D{mdata}:Budget!D{mdata+11}),"
            f"SUMIF(Budget!A{mdata}:Budget!A{mdata+11},C3,Budget!D{mdata}:Budget!D{mdata+11})),0)"
        )

    if inv_ws:
        ws.cell(row=6, column=7).value = "=IFERROR(SUM(Investments!E4:E53),0)"

    for col in kpi_cols:
        ws.cell(row=6, column=col).number_format = cfmt
        ws.row_dimensions[6].height = 32

    # Charts pulled from other sheets
    CHART_ROW = "A9"

    if budget_ws and bud_rows:
        mhead, mdata = bud_rows
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Monthly Income vs Expenses"
        chart.width = 18; chart.height = 12
        inc_r = Reference(budget_ws, min_col=2, min_row=mhead, max_row=mdata + 11)
        exp_r = Reference(budget_ws, min_col=3, min_row=mhead, max_row=mdata + 11)
        cats_r = Reference(budget_ws, min_col=1, min_row=mdata, max_row=mdata + 11)
        chart.add_data(inc_r, titles_from_data=True)
        chart.add_data(exp_r, titles_from_data=True)
        chart.set_categories(cats_r)
        ws.add_chart(chart, "A9")

    if inv_ws:
        chart2 = DoughnutChart()
        chart2.title = "Portfolio Allocation"
        chart2.width = 14; chart2.height = 12
        inv_labels = Reference(inv_ws, min_col=1, min_row=4, max_row=4 + 9)
        inv_vals   = Reference(inv_ws, min_col=5, min_row=3, max_row=4 + 9)
        chart2.add_data(inv_vals, titles_from_data=True)
        chart2.set_categories(inv_labels)
        ws.add_chart(chart2, "J9")

    set_col_widths(ws, [16, 4, 16, 4, 16, 4, 16, 4, 4, 4])
    ws.sheet_view.showGridLines = False

# ── All-in-one sheet ──────────────────────────────────────────────────────────

def build_allinone(ws, setup, mods, theme, cfmt,
                   budget=None, investments=None, goals_data=None, business=None):
    sym = setup["cur"][1]
    row = [1]  # mutable row counter

    def cur_row():
        return row[0]

    def next_row(n=1):
        row[0] += n
        return row[0]

    def section_title(title, ncols=8):
        r = cur_row()
        merge_title(ws, title, r, 1, ncols, theme)
        next_row(2)

    section_title("Financial Overview — All-in-One", 8)

    # KPI summary
    kpis = ["Total Income", "Total Expenses", "Net Savings"]
    for i, k in enumerate(kpis):
        c = i * 2 + 1
        ws.cell(row=cur_row(), column=c, value=k).font = _font(theme["kpi_title"], bold=True)
        ws.cell(row=cur_row(), column=c).fill = _fill(theme["header_bg"])
        ws.cell(row=cur_row(), column=c).alignment = _align()
        ws.cell(row=cur_row(), column=c + 1, value=0).number_format = cfmt
        ws.cell(row=cur_row(), column=c + 1).font = _font(theme["accent"], bold=True, size=14)
        ws.cell(row=cur_row(), column=c + 1).fill = _fill(theme["bg"])
        ws.cell(row=cur_row(), column=c + 1).alignment = _align()
    next_row(2)

    if "Personal Budget" in mods and budget:
        section_title("Personal Budget — Weekly Tracker", 6)
        cats = budget["categories"]
        headers = ["Week", "Week Start", "Income"] + cats + ["Total Exp", "Net"]
        nc = len(headers)
        style_row(ws, cur_row(), theme, nc, kind="header")
        for ci, h in enumerate(headers, 1):
            ws.cell(row=cur_row(), column=ci, value=h)
        next_row()
        today = datetime.date.today()
        year_start = datetime.date(today.year, 1, 4)
        while year_start.weekday() != 0:
            year_start += datetime.timedelta(days=1)
        cur_week = today.isocalendar()[1]
        ec = 3 + len(cats)
        tc = ec + 1
        nc2 = tc + 1
        for w in range(1, 53):
            r = cur_row()
            style_row(ws, r, theme, nc, alt=(w % 2 == 0))
            wd = year_start + datetime.timedelta(weeks=w - 1)
            ws.cell(row=r, column=1, value=w)
            ws.cell(row=r, column=2, value=wd).number_format = "DD MMM YYYY"
            ws.cell(row=r, column=3, value=budget["weekly_income"] if w == cur_week else 0)
            for ci in range(4, ec + 1):
                ws.cell(row=r, column=ci, value=0)
            exp_r = f"{get_column_letter(4)}{r}:{get_column_letter(ec)}{r}"
            ws.cell(row=r, column=tc, value=f"=SUM({exp_r})")
            ws.cell(row=r, column=nc2, value=f"=C{r}-{get_column_letter(tc)}{r}")
            apply_money_fmt_range(ws, r, r, [3] + list(range(4, nc2 + 1)), cfmt)
            next_row()
        next_row()

    if "Investments" in mods and investments:
        section_title("Investments — Stocks & ETFs", 7)
        style_row(ws, cur_row(), theme, 7, kind="header")
        for ci, h in enumerate(["Ticker", "Buy Price", "Qty", "Current Price", "Total Value", "P&L", "% Change"], 1):
            ws.cell(row=cur_row(), column=ci, value=h)
        next_row()
        for s in investments["stocks"]:
            r = cur_row()
            style_row(ws, r, theme, 7)
            ws.cell(row=r, column=1, value=s["ticker"])
            ws.cell(row=r, column=2, value=s["buy"]).number_format = cfmt
            ws.cell(row=r, column=3, value=s["qty"])
            ws.cell(row=r, column=4, value=s["cur"]).number_format = cfmt
            ws.cell(row=r, column=5, value=f"=C{r}*D{r}").number_format = cfmt
            ws.cell(row=r, column=6, value=f"=E{r}-(B{r}*C{r})").number_format = cfmt
            ws.cell(row=r, column=7, value=f"=IF(B{r}*C{r}=0,0,(E{r}-(B{r}*C{r}))/(B{r}*C{r}))").number_format = "0.00%"
            next_row()
        next_row()

    if "Savings Goals" in mods and goals_data:
        section_title("Savings Goals", 7)
        style_row(ws, cur_row(), theme, 7, kind="header")
        for ci, h in enumerate(["Goal", "Target", "Saved", "Monthly", "Months Left", "% Done", "Progress"], 1):
            ws.cell(row=cur_row(), column=ci, value=h)
        next_row()
        g_start = cur_row()
        for g in goals_data["goals"]:
            r = cur_row()
            style_row(ws, r, theme, 7)
            ws.cell(row=r, column=1, value=g["name"])
            ws.cell(row=r, column=2, value=g["target"]).number_format = cfmt
            ws.cell(row=r, column=3, value=g["current"]).number_format = cfmt
            ws.cell(row=r, column=4, value=g["monthly"]).number_format = cfmt
            ws.cell(row=r, column=5, value=f"=IF(D{r}=0,\"—\",ROUNDUP((B{r}-C{r})/D{r},0))")
            ws.cell(row=r, column=6, value=f"=IF(B{r}=0,0,C{r}/B{r})").number_format = "0.00%"
            next_row()
        g_end = cur_row() - 1
        db = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                         color=theme["accent"])
        ws.conditional_formatting.add(f"F{g_start}:F{g_end}", db)
        next_row()

    if "Business / Freelance" in mods and business:
        section_title("Business / Freelance — Invoices", 6)
        style_row(ws, cur_row(), theme, 6, kind="header")
        for ci, h in enumerate(["Client", "Date", "Invoice #", "Service", "Amount", "Status"], 1):
            ws.cell(row=cur_row(), column=ci, value=h)
        next_row()
        inv_start = cur_row()
        for ri in range(20):
            r = cur_row()
            style_row(ws, r, theme, 6, alt=(ri % 2 == 0))
            ws.cell(row=r, column=2).number_format = "DD MMM YYYY"
            ws.cell(row=r, column=5).number_format = cfmt
            next_row()
        inv_end = cur_row() - 1
        dv_s = DataValidation(type="list", formula1='"Paid,Unpaid,Overdue"', showDropDown=False)
        ws.add_data_validation(dv_s)
        for ri in range(inv_start, inv_end + 1):
            dv_s.add(ws.cell(row=ri, column=6))
        next_row()

    set_col_widths(ws, [18, 14, 14, 14, 14, 14, 14, 14])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ── Workbook builder ──────────────────────────────────────────────────────────

def create_workbook(setup, mods, budget, investments, goals_data, business):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    theme = setup["theme"]
    cfmt  = setup["cur"][2]

    if setup["multi"]:
        budget_ws = inv_ws = goals_ws = biz_ws = None
        bud_rows  = None

        if "Personal Budget" in mods and budget:
            budget_ws = wb.create_sheet("Budget")
            bud_rows  = build_budget(budget_ws, budget, theme, cfmt)

        if "Investments" in mods and investments:
            inv_ws = wb.create_sheet("Investments")
            build_investments(inv_ws, investments, theme, cfmt)

        if "Savings Goals" in mods and goals_data:
            goals_ws = wb.create_sheet("Savings Goals")
            build_savings(goals_ws, goals_data, theme, cfmt)

        if "Business / Freelance" in mods and business:
            biz_ws = wb.create_sheet("Business")
            build_business(biz_ws, business, theme, cfmt)

        dash_ws = wb.create_sheet("Dashboard", 0)
        build_dashboard(dash_ws, setup, mods, budget_ws, inv_ws, goals_ws, biz_ws,
                        bud_rows, theme, cfmt)
    else:
        ws = wb.create_sheet("Finance Tracker")
        build_allinone(ws, setup, mods, theme, cfmt, budget, investments, goals_data, business)

    return wb


# ── VBA injection ─────────────────────────────────────────────────────────────

VBA_MODULE = """
Sub RefreshDashboard()
    Application.CalculateFull
    Dim obj As Object
    For Each obj In ActiveSheet.ChartObjects
        obj.Chart.Refresh
    Next obj
    MsgBox "Dashboard refreshed!", vbInformation, "Finance Tracker"
End Sub
"""

VBA_SHEET_EVENT = """
Private Sub Worksheet_Change(ByVal Target As Range)
    If Not Intersect(Target, Me.Range("C3")) Is Nothing Then
        Application.CalculateFull
    End If
End Sub
"""

def hex_to_rgb(h):
    h = h.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return r + g * 256 + b * 65536  # Excel BGR-as-int

def inject_vba(xlsx_path, xlsm_path, theme, multi_sheet):
    try:
        import win32com.client as win32
    except ImportError:
        console.print("[yellow]pywin32 not available — skipping VBA injection.[/yellow]")
        return False

    console.print("[dim]Opening Excel to inject VBA macros...[/dim]")
    excel = None
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(str(pathlib.Path(xlsx_path).resolve()))

        try:
            vb = wb.VBProject
        except Exception:
            console.print(
                "[yellow]VBA injection blocked by Excel Trust Center.\n"
                "Enable: Excel → Options → Trust Center → Trust Center Settings\n"
                "→ Macro Settings → ☑ Trust access to the VBA project object model[/yellow]"
            )
            wb.Close(False)
            return False

        # Add standard module
        mod = vb.VBComponents.Add(1)  # vbext_ct_StdModule
        mod.Name = "FinanceMacros"
        mod.CodeModule.AddFromString(VBA_MODULE)

        # Add Worksheet_Change to Dashboard (or first sheet)
        target_name = "Dashboard" if multi_sheet else "Finance Tracker"
        for i in range(1, wb.Sheets.Count + 1):
            sh = wb.Sheets(i)
            if sh.Name == target_name:
                comp = vb.VBComponents(sh.CodeName)
                comp.CodeModule.AddFromString(VBA_SHEET_EVENT)

                # Add Refresh button
                btn = sh.Shapes.AddShape(
                    1,      # msoShapeRectangle
                    wb.Sheets(i).UsedRange.Width - 170,
                    8, 160, 28
                )
                btn.Name = "btn_refresh"
                btn.OnAction = "FinanceMacros.RefreshDashboard"
                btn.Fill.ForeColor.RGB = hex_to_rgb(theme["accent"])
                btn.Line.Visible = False
                btn.TextFrame.Characters().Text = "  Refresh Dashboard"
                btn.TextFrame.Font.Bold = True
                btn.TextFrame.Font.Color = hex_to_rgb("FFFFFF")
                btn.TextFrame.Font.Size = 10
                break

        wb.SaveAs(str(pathlib.Path(xlsm_path).resolve()), FileFormat=52)
        wb.Close(False)
        return True

    except Exception as e:
        console.print(f"[yellow]VBA injection error: {e}[/yellow]")
        return False
    finally:
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    console.clear()
    console.print(Panel(
        "[bold cyan]EXCEL FINANCE GENERATOR[/bold cyan]\n"
        "[dim]Generates a personalised, macro-enabled Excel spreadsheet[/dim]",
        padding=(1, 8), style="cyan"
    ))

    # Step 1-6: Collect inputs
    setup = collect_setup()
    mods  = collect_modules()

    sym = setup["cur"][1]
    budget      = collect_budget(sym)      if "Personal Budget"     in mods else None
    investments = collect_investments(sym) if "Investments"         in mods else None
    goals_data  = collect_goals(sym)       if "Savings Goals"       in mods else None
    business    = collect_business(sym)    if "Business / Freelance" in mods else None

    # Confirm
    console.print()
    console.print(Panel(
        f"[bold]Ready to generate[/bold]\n"
        f"File: [cyan]{setup['fname']}.xlsm[/cyan]\n"
        f"Theme: [cyan]{setup['th_key']}[/cyan]  |  Currency: [cyan]{setup['cur_key']}[/cyan]\n"
        f"Modules: [cyan]{', '.join(mods)}[/cyan]",
        style="green", padding=(0, 4)
    ))
    if not ok("Generate spreadsheet now?"):
        console.print("[yellow]Cancelled.[/yellow]")
        return

    # Log answers
    log_answers(setup, mods, budget, investments, goals_data, business)
    console.print("[dim]Answers logged to answers_log.txt[/dim]")

    # Build workbook
    console.print("[dim]Building Excel workbook...[/dim]")
    wb = create_workbook(setup, mods, budget, investments, goals_data, business)

    # Save paths
    OUTPUT_DIR.mkdir(exist_ok=True)
    today_str = datetime.date.today().strftime("%Y%m%d")
    base_name = f"{setup['fname']}_{today_str}"
    xlsx_path = OUTPUT_DIR / f"{base_name}.xlsx"
    xlsm_path = OUTPUT_DIR / f"{base_name}.xlsm"

    wb.save(str(xlsx_path))
    console.print(f"[dim]Saved interim .xlsx[/dim]")

    # VBA injection
    vba_ok = inject_vba(str(xlsx_path), str(xlsm_path), setup["theme"], setup["multi"])

    if vba_ok:
        os.remove(str(xlsx_path))
        final_path = xlsm_path
    else:
        final_path = xlsx_path
        console.print("[yellow]Saved as .xlsx (no VBA). Open manually to enable macros.[/yellow]")

    console.print()
    console.print(Panel(
        f"[bold green]Done![/bold green]\n"
        f"[cyan]{final_path}[/cyan]",
        style="green", padding=(0, 4)
    ))

    # Open it
    if ok("Open the file now?", default=True):
        os.startfile(str(final_path))


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n[yellow]Aborted.[/yellow]")
        sys.exit(0)
